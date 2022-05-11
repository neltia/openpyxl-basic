# 4_datas_read.py
from openpyxl import load_workbook

wb = load_workbook('wb_multiplication.xlsx')
ws = wb['Sheet']

print(ws["C"])
for cell in ws["C"]:
    print(cell.value)

print(ws["2"])
for cell in ws["2"]:
    print(cell.value)

print(" ".join([cell.value for cell in ws["2"] if cell.value is not None]))
for row in range(1, 9):
    for column in range(1, 6):
        c = ws.cell(row=row+2, column=column+1)
        print("%3d" %(c.value), end=' ')
    print()

for row in ws.iter_rows(min_row=2, max_row=9, min_col=2, max_col=6):
    for cell in row:
        print(cell.value, end=' ')
    print()


