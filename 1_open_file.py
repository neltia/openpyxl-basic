# 1_open_file.py
from openpyxl import load_workbook

# - 워크북 불러오기
wb = load_workbook('wb.xlsx')
# - 워크시트 이름 목록 출력
print(wb.sheetnames)
