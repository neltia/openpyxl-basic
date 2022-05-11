# 3_datas_insert.py
from openpyxl import Workbook
wb = Workbook()

ws = wb.active # 기본 시트 가져오기

# 첫째행 타이틀 적기 예제
# 제목 적기
sub = ['번호', '이름', '주소', '이메일']
for kwd, j in zip(sub, range(1, len(sub)+1)):
    ws.cell(row=2, column=j+1).value = kwd

# wb.save("wb_subtitle.xlsx")
# import os; os.remove("wb_subtitle.xlsx")

# 구구단
# - 구구단 단 입력 (1 ~ 5단)
for i in range(1, 6):
    ws.cell(row=2, column=i+1).value = f"{i}단"

# - 내용 입력
for row in range(1, 10):
    for column in range(1, 6):
        # 초기화하면서 데이터 입력
        c = ws.cell(row=row+2, column=column+1, value=row*column)
        # 데이터 수정
        c.value=row*column

wb.save("wb_multiplication.xlsx")


