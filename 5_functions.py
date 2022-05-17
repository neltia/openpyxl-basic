# 5_functions.py
# - 구구단 숫자 표에서 기초 통계를 구합니다.
from openpyxl import load_workbook

wb = load_workbook('wb_multiplication.xlsx')
ws = wb['Sheet']

# 구구단 헤드 지정 및 n단까지 있는지 확인
wrow = ws["2"]
wrow_data = [cell.value for cell in wrow if cell.value is not None]

# 사용할 통계 함수
func_table = {
    '합계': "SUM",
    "평균": "AVERAGE",
    "최댓값": "MAX",
    "최솟값": "MIN"
}

# 사용할 통계 텍스트 입력
start_row = 11
sub = list(func_table.keys())
for kwd, j in zip(sub, range(1, len(sub)+1)):
    ws.cell(row=start_row+j, column=1).value = kwd

# 통계 함수 적용
# - 테스트용
ws["B12"].value = "=SUM(B3:B11)"
# - 함수 입력 자동화
func_list = list(func_table.values())
for row in range(1, len(sub)+1):
    func = func_list[row-1]

    for column in range(1, len(wrow_data)+1):
        work_col = chr(65+column) # 65=A
        formula_range = f"{work_col}3:{work_col}11"
        ws[start_row+row][column].value = f"={func}({formula_range})"

# 결과물 저장
wb.save("wb_mutlple_func.xlsx")
