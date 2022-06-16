# 12_cell_border.py
# - Post 10. 셀 테두리와 배경색

from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Border, Side
from copy import copy

# 5_functions.py 파일의 결과 파일 경로 입력 후 파일 load
path = "./preview_resultfile"
wb = load_workbook(f'{path}/wb_multiple_func.xlsx')
ws = wb['Sheet']

# 셀 범위 이동
ws.move_range("A2:F15", cols=1, translate=True)

# 데이터 정렬
# - B2행 가운데 정렬
cell_list = [cell for cell in ws["2"] if cell is not None]
for cell in cell_list:
    cell.alignment = Alignment(horizontal='center')
# - B12:B15 셀 가운데 정렬
for cell in range(12, 16):
    ws[f"B{cell}"].alignment = Alignment(horizontal='center')

# 테두리 적용
# - Border 인자 값 확인
print(dir(Border))
# - 일부 셀에 적용 테스트
ws["C2"].border = Border(
    top=Side(border_style="thick", color='000000')
)
ws["E5"].border = Border(Side(border_style="double", color='F15F5F'))
ws["D7"].border = Border(
    left=Side(border_style="thin"),
    right=Side(border_style="thin"),
)

# - 셀 기본 테두리 설정
bd_thin = Side(border_style='thin', color='000000')
THIN_BORDER = Border(bd_thin, bd_thin, bd_thin, bd_thin)
for rng in ws["B2:G15"]:
    for cell in rng:
        cell.border = THIN_BORDER

# - 특정 범위 굵은 바깥 테두리 설정
# -- 스타일 정의
bd_thick = Side(border_style='thick', color='000000')
THICK_BORDER = Border(
    top=bd_thick, bottom=bd_thick, left=bd_thick, right=bd_thick
)
THICK_Top = Border(
    top=bd_thick, bottom=bd_thin, left=bd_thin, right=bd_thin
)
THICK_Bottom = Border(
    top=bd_thin, bottom=bd_thick, left=bd_thin, right=bd_thin
)
THICK_Left = Border(
    top=bd_thin, bottom=bd_thin, left=bd_thick, right=bd_thin
)
THICK_Right = Border(
    top=bd_thin, bottom=bd_thin, left=bd_thin, right=bd_thick
)
borderTopLeft = Border(
    top=bd_thick, bottom=bd_thin, left=bd_thick, right=bd_thin
)
borderTopRight = Border(
    top=bd_thick, bottom=bd_thin, left=bd_thin, right=bd_thick
)
borderBottomLeft = Border(
    top=bd_thin, bottom=bd_thick, left=bd_thick, right=bd_thin
)
borderBottomRight = Border(
    top=bd_thin, bottom=bd_thick, left=bd_thin, right=bd_thick
)
borderSide = Side(border_style='thick') # applies to sides of each cell

rowTop = 2
rowBot = 15
colLeft = 2
colRight = 7
rows = range(rowTop, rowBot+1)
cols = range(colLeft, colRight+1)
start_cell = chr(64 + rowTop)
end_cell = chr(64 + colRight)

# -- 범위 내 셀들을 ws.cell 지정 방식으로 지정해 기본 테두리 설정
for row in rows:
    for col in cols:
        ws.cell(row, col).border = THIN_BORDER

# -- 바깥쪽 테두리 굵게 설정
for row in rows:
    for col in cols:
        ws.cell(rowTop, col).border = THICK_Top
        ws.cell(rowBot, col).border = THICK_Bottom
        ws.cell(row, colLeft).border = THICK_Left
        ws.cell(row, colRight).border = THICK_Right
        ws[f'{start_cell}{rowTop}'].border = borderTopLeft
        ws[f'{end_cell}{rowTop}'].border = borderTopRight
        ws[f'{start_cell}{rowBot}'].border = borderBottomLeft
        ws[f'{end_cell}{rowBot}'].border = borderBottomRight

# - 특정 행 굵은 바깥 테두리 설정
for rng in ws["B2:G2"]:
    for idx, cell in enumerate(rng):
        if idx == 0:
            cell.border = Border(top=bd_thick, bottom=bd_thick, left=bd_thick, right=bd_thin)
        elif idx == len(rng)-1:
            cell.border = Border(top=bd_thick, bottom=bd_thick, left=bd_thin, right=bd_thick)
        else:
            cell.border = Border(top=bd_thick, bottom=bd_thick, left=bd_thin, right=bd_thin)
# - 서식 복사를 활용해 테두리 설정
for col in cols:
    about_cell = ws.cell(15, col)               # 스타일을 가져올 셀
    copy_cell = ws.cell(11, col)                # 스타일 복사가 수행될 셀
    copy_cell._style = copy(about_cell._style)  # 스타일 전체 복사

# 작업 내용 저장
wb.save("wb_border.xlsx")
