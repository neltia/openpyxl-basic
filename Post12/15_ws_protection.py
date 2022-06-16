# 15_ws_protection.py
# - Post 12. 시트 설정

from openpyxl import load_workbook
from openpyxl.styles import Protection

# 12_cell_border.py 파일의 결과 파일 경로 입력 후 파일 load
path = "../Post10"
wb = load_workbook(f'{path}/wb_border.xlsx')
ws = wb['Sheet']

# 구구단 표 데이터만 수정을 허락하도록 설정
for row in ws["B2:G15"]:
    for cell in row:
        cell.protection = Protection(locked=False)

# protection 설정
ws.protection.password = "*****"
ws.protection.enable()

# 설정을 적용한 파일 저장
wb.save("wb_protection.xlsx")
