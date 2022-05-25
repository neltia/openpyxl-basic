# 13_cell_bgcolor.py
# - Post 10. 셀 테두리와 배경색

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Color
from copy import copy

# 12_cell_border.py 파일의 결과 파일 경로 입력 후 파일 load
wb = load_workbook(f'wb_border.xlsx')
ws = wb['Sheet']

# 색 설정
c2 = ws["C2"]
c2.fill = PatternFill(fill_type='solid',
                      fgColor=Color('FDE9D9'))

# - 서식 복사를 활용해 배경색 설정
for col in range(4, 8):
    ws.cell(2, col).fill = copy(c2.fill)

# 작업 내용 저장
wb.save("wb_cellfill.xlsx")
