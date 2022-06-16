# 6_display_formatting.py

from openpyxl import Workbook
import datetime

wb = Workbook()
ws = wb.active

# 숫자 형식 적용
# - 세 자리마다 쉼표 표기
value = 300000
ws['B2'] = value
ws['B2'].number_format = '#,##0'
# - 조건 형식 설정
ws['C2'] = value
ws['C3'] = 0
ws['C4'] = -value
for rng in ws["C2:C4"]:
    for cell in  rng:
        cell.number_format = '[RED]#,##0;[BLUE]-#,##0;"-"'
# - 적용 결과 확인
# wb.save("test.xlsx")

# 날짜 헝식 적용
# - 기본 날짜 데이터 입력
ws["E1"] = "적용 전"
ws['E2'] = "2022-05-17"                     # 텍스트로 날짜 입력
ws['E3'] = datetime.datetime.now()          # 현재 시간 입력
ws['E4'] = datetime.datetime(2022, 5, 17)   # 특정 시간 입력
# - 입력된 데이터의 기본 셀 서식 형식
print(ws['E2'].number_format)               # General = G/표준
print(ws['E3'].number_format)               # yyyy-mm-dd h:mm:ss
# - 입력 결과 확인
# wb.save("test.xlsx")
# - 날짜 서식 적용
ws['E2'].number_format = 'yyyy년 mm월 dd일 ddd (aaa)'
ws['E3'].number_format = 'yyyy년 mm월 dd일 ddd (aaa)'

# 결과물 저장
wb.save("wb_formatting.xlsx")
