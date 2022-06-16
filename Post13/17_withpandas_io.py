# 17_withpandas_io.py
# - Post 13. 판다스(pandas)와 같이 사용하기

# 라이브러리 호출
from openpyxl import Workbook
from openpyxl import load_workbook
import openpyxl.utils.dataframe
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd

"""
dataframe to openpyxl
"""
# 데이터프레임 기반 dict 데이터 선언
data = {
    'id' : [1,2,3],
    'name' : ['Choi','Jeong','Kim'],
    'age' : [10,20,30],
    'assets' : [150.4, 123.4, 88.88],
    'job' : ['Student', 'CEO', 'Dad']
}
# 데이터프레임 변환
df = pd.DataFrame(data)

# 데이터프레임을 엑셀 파일로 통째로 추출
# - 기본 사용
df.to_excel("new_file.xlsx")
# - 시트 이름 지정
# df.to_excel("new_file.xlsx", sheet_name='Sheet_name_1')
# - 인덱스 제외 설정
# df.to_excel("new_file.xlsx", index=False)

# 데이터프레임 유틸리티 지원 함수 목록 확인
print(dir(openpyxl.utils.dataframe))

# 새 워크시트 생성
wb = Workbook()
ws = wb.active

# 열 단위 데이터 추가(append)
for r in dataframe_to_rows(df, index=True, header=True):
    ws.append(r)

# 스타일 지정
for cell in ws['A'] + ws[1]:
    cell.style = 'Pandas'

# 결과 확인
wb.save("wb_pandas_openpyxl.xlsx")

# 파일 불러와 삽입
wb = load_workbook('wb_pandas_openpyxl.xlsx')
ws = wb["Sheet"]
for r in dataframe_to_rows(df, index=False, header=True):
    ws.append(r)
wb.save("wb_pandas_openpyxl.xlsx")

# 계산
print("누적 합")
for calc in openpyxl.utils.dataframe.accumulate(df["assets"]):
    print(calc)
print("곱 계산")
print(openpyxl.utils.dataframe.prod(df["age"]))
print("누적 곱")
for calc in openpyxl.utils.dataframe.accumulate(df["assets"], openpyxl.utils.dataframe.operator.mul):
    print(calc)
print("누적 차")
# print(dir(openpyxl.utils.dataframe.operator))

"""
openpyxl to dataframe
"""
# 첫 번째 방법
df = pd.read_excel("wb_pandas_openpyxl.xlsx")
print(df)

# 두 번째 방법
df = pd.DataFrame(ws.values)
# print(df)
df.columns = df.iloc[0, :]
df = df.iloc[1:, :]
print(df)
