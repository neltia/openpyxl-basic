# 11_ws_movecells.py
# - Post 9. 데이터 잘라내 붙이기

from openpyxl import load_workbook
from openpyxl.styles import Alignment

# 5_functions.py 파일의 결과 파일 경로 입력 후 파일 load
path = "../Post04"
wb = load_workbook(f'{path}/wb_multiple_func.xlsx')
ws = wb['Sheet']

# move_range 수행
ws.move_range("A2:F15", rows=3, cols=1)
wb.save('wb_moving.xlsx')

wb = load_workbook(f'{path}/wb_multiple_func.xlsx')
ws = wb['Sheet']
ws.move_range("A2:F15", rows=3, cols=1, translate=True)
wb.save('wb_moving2.xlsx')
