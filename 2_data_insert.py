# 2_data_insert.py
from openpyxl import load_workbook

wb = load_workbook('wb.xlsx')
ws = wb['Sheet']

# 셀 값 설정 및 가져오기
print(ws['A1'].value)
ws['A1'] = "TEST DATA"
print(ws['A1'].value)

# 셀 값을 가져오는 다른 방법
print(ws.cell(row=1, column=1))

# 작업 내용 저장
wb.save('wb_insert.xlsx')

# print(type(ws.cell(row=1, column=1)))
# print(dir(ws.cell(row=1, column=1)))
