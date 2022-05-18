# 8_cell_font.py
# - Post 6. 정렬과 글꼴
# - URL:

from openpyxl import Workbook
from openpyxl.styles import Font

wb = Workbook()
ws = wb.active

# 셀 변수 설정
b2 = ws["B2"]
c3 = ws["C3"]
d4 = ws["d4"]

b2.value = "글자 테스트"
d4.value = "글자 테스트"
c3.value = "글자 테스트"

# 폰트 설정
b2.font = Font(name='맑은 고딕',                # 폰트 이름으로 폰트 설정
               size=11,                        # 글자 크기
               bold=True,                      # 굵게 설정 여부 (Default: False)
               italic=True,                    # 기울임 여부 (Default: False)
               vertAlign='subscript',          # 첨자
               underline='singleAccounting',   # 밑줄
               strike=True,                    # 취소선
               color='0000FF00')               # 폰트 색상

c3.font = Font(name='나눔고딕',
               size=10,
               vertAlign='baseline',
               color='000000FF')

d4.font = Font(name='궁서',
               size=15,
               strike=True,
               underline='double',
               vertAlign='superscript',
               color='FF0000FF')

wb.save("wb_font.xlsx")
