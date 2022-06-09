# 16_ws_heiwid.py
# - Post 12. 시트 설정

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# 12_cell_border.py 파일의 결과 파일 경로 입력 후 파일 load
path = "./preview_resultfile"
wb = load_workbook(f'{path}/wb_border.xlsx')
ws = wb['Sheet']

# 작업 시트 복사
ws2 = wb.copy_worksheet(ws)
# print(ws2.title)
ws2.title = "Sheet2" # 복사한 시트의 기존 이름은 "<기존 시트 이름> Copy"
ws3 = wb.copy_worksheet(ws)
ws3.title = "Sheet3"

# 행 높이 설정
ws.row_dimensions[1].height = 10
for row in range(2, 10):
    ws.row_dimensions[row].height = 50

# 열 너비 설정
ws2.column_dimensions["A"].width  = 1
# for col in range(2, 10):
#     ws2.column_dimensions[chr(64 + col)].width  = 15
for col in range(2, 10):
    ws2.column_dimensions[get_column_letter(col)].width  = 20

# 열 너비 자동 맞춤 설정
ws3["G5"] = "1234567891012345678910"


def AutoFitColumnSize(worksheet, columns=None, margin=2):
    for i, column_cells in enumerate(worksheet.columns):
        is_ok = False
        if columns == None:
            is_ok = True
        elif isinstance(columns, list) and i in columns:
            is_ok = True

        if is_ok:
            length = max(len(str(cell.value)) for cell in column_cells)
            worksheet.column_dimensions[column_cells[0].column_letter].width = length + margin
    return worksheet


AutoFitColumnSize(ws3)

# 파일 저장
wb.save("wb_rowcol_sizing.xlsx")