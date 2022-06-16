# 7_cell_alignment.py
# - Post 6. 정렬과 글꼴
# - URL: https://blog.naver.com/dsz08082/222737385550

from openpyxl import load_workbook
from openpyxl.styles import Alignment

# 5_functions.py 파일의 결과 파일 경로 입력 후 파일 load
path = "../Post04"
wb = load_workbook(f'{path}/wb_multiple_func.xlsx')
ws = wb['Sheet']

# B2행 가운데 정렬
cell_list = [cell for cell in ws["2"] if cell is not None]
for cell in cell_list:
    cell.alignment = Alignment(horizontal='center')

wb.save("wb_alignment.xlsx")
