# 9_cell_merge.py
# - Post 7. 셀 병합/분리

from openpyxl import load_workbook
from openpyxl.styles import Alignment

# 5_functions.py 파일의 결과 파일 경로 입력 후 파일 load
path = "../Post04"
wb = load_workbook(f'{path}/wb_multiple_func.xlsx')
ws = wb['Sheet']

# 셀 병합
# - 헤더 병합
ws["B1"] = "구구단"
ws.merge_cells("B1:F1")
# - 왼쪽 상단 데이터만 남는 것 확인
ws.merge_cells("B2:F2")
# - 셀의 위치를 사용한 셀 병합
ws.merge_cells(start_row=12, start_column=1, end_row=13, end_column=1)
ws.merge_cells(start_row=7, start_column=7, end_row=10, end_column=9)

# - 작업 내용 저장
wb.save("wb_merge.xlsx")

# 셀 병합 해제
ws["B1"].alignment = Alignment(horizontal='center')
ws.unmerge_cells("B2:F2")
ws.unmerge_cells(start_row=12, start_column=1, end_row=13, end_column=1)
ws.unmerge_cells(start_row=7, start_column=7, end_row=10, end_column=9)

wb.save("wb_unmerge.xlsx")
