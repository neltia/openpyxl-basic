"""
220110_B_works1.py
- Post 14. 응용 사례 - ITQ 엑셀 문제 풀이
[22년 01월 10일 기출문제 복원본 풀이] 중 제2작업
- 문제 출처 : https://www.comcbt.com/xe/itqe/5563038
"""

# 필요 라이브러리 호출
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.styles import Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
from copy import copy
import time

# 제1작업 시트 불러오기
wb = load_workbook('result-2201010B-openpyxl_part1.xlsx')
ws1 = wb['제1작업']
ws2 = wb.create_sheet("제2작업")

# A열 너비 조정
ws2.column_dimensions["A"].width = 1

# 셀 기본 스타일 지정 함수
def cell_style(workseet, cell, align=True, thin=True):
    cell.font = Font(name='굴림', size=11)
    if align:
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    if thin:
        cell.border = THIN_BORDER
    return workseet

# 기존 데이터를 다른 시트로 복사
min_row, max_row = 4, 12
min_col, max_col = 2, 8
# copying the cell values from source
# excel file to destination excel file
for i in range(min_row, max_row+1):
    for j in range (min_col, max_col+1):
        # reading cell value from source excel file
        origin_cell = ws1.cell(row=i, column=j)
        # writing the read value to destination excel file
        cell = ws2.cell(row=i-2, column=j)
        cell.value = origin_cell.value
        cell._style = copy(origin_cell._style)

# 제1작업 시트 삭제
del wb['제1작업']

# 테두리 정의
bd_thin = Side(border_style='thin')
THIN_BORDER = Border(bd_thin, bd_thin, bd_thin, bd_thin)

# B11, H11 - 급여 전체 평균
b11 = ws2["B11"]
b11.value = "급여(단위:원) 전체 평균"
cell_style(ws2, b11)
ws2.merge_cells("B11:G11")
h11 = ws2["H11"]
h11.value = "=AVERAGE(H3:H10)"
h11.number_format = '0,000'
cell_style(ws2, h11)

# 목표값 찾기
# - 수식 셀: H11
# - 찾는 값: 3200000
# - 값을 바꿀 셀: $H$3 (원본 값 변경)
# *** 목표값 칮기 기능 자체는 openpyxl의 미지원 기능으로 판단 ***
# 대안: 다음으로 해당 기능 수행 결과 대체 가능
# 변경대상값 = 기존 평균값에서 변경할 평균 값 * 총 자료수 - 대상값을 제외한 합
# 23006000 + x / 8 = 3200000 => x = 3200000 * 8 - 23006000
nums = 0
for row in ws2["H4:H10"]:
    for cell in row:
        nums += cell.value
ws2["H11"].value = 3200000
ws2["H3"].value = 3200000 * 8 - nums

# 고급필터
# 발령부서가 '배송부'이거나, 근속기간이 '2'이하인 자료 이름, 발령구분, 근속기간, 급여 데이터만 추출
# - 조건 범위: B14 셀부터 입력
# - 복사 위치: B18 셀부터 나타나도록
# *** 고급필터 기능 자체는 openpyxl의 미지원 기능으로 판단 ***
# 자동필터 기능 지원 테스트
sub = ['사원코드', '이름', '발령부서', '발령구분', '근속기간', '출생년', '급여\n(단위:원)']
ws_test = wb.copy_worksheet(wb["제2작업"])
ws_test.title = "자동 필터 테스트"
ws_test.auto_filter.ref = "B2:H10"
ws_test.auto_filter.add_filter_column(0, sub)

# 대안: 파이썬 판다스 필터 사용
# - filter 조건
# -> ws2[B14], ws2[B15]
cell = ws2["B14"]
cell.value = "발령부서"
cell._style = copy(ws2["D2"]._style)
cell = ws2["B15"]
cell.value = "배송부"
cell.alignment = Alignment(horizontal='left')
cell_style(ws2, cell, False, False)
# -> ws2[C14], ws2[C15]
cell = ws2["C14"]
cell.value = "근속기간"
cell._style = copy(ws2["F2"]._style)
cell = ws2["C16"]
cell.value = "<=2"
cell.alignment = Alignment(horizontal='left')
cell_style(ws2, cell, False, False)

# - 필터 데이터 적용
df = pd.DataFrame(ws2.values)
df = df.drop(0, axis=1)
df.columns = df.iloc[1, :]
df = df[2:-1]
mask1 = (df.발령부서 == "배송부") | (df.근속기간 <= 2)
df_filter = df.loc[mask1,:]

for r in dataframe_to_rows(df_filter, index=True, header=True):
   ws2.append(r)

# 완료 데이터 저장
wb.save("result-2201010B-openpyxl_part2.xlsx")
