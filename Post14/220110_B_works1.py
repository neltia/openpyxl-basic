"""
220110_B_works1.py
- Post 14. 응용 사례 - ITQ 엑셀 문제 풀이
[22년 01월 10일 기출문제 복원본 풀이] 중 제1작업
- 문제 출처 : https://www.comcbt.com/xe/itqe/5563038
"""

# 필요 라이브러리 호출
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.styles import PatternFill, Color
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting import Rule
from openpyxl.styles.differential import DifferentialStyle
from copy import copy

# 새 워크북 생성 및 시트 이름 변경
wb = Workbook()
ws = wb.active
ws.title = "제1작업"

# A열 너비 조정
ws.column_dimensions["A"].width = 1

# 셀 기본 스타일 지정 함수
def cell_style(workseet, cell):
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.font = Font(name='굴림', size=11)
    return workseet

# 데이터 입력
# - 4행 데이터 입력
sub = ['사원코드', '이름', '발령부서', '발령구분', '근속기간', '출생년', '급여\n(단위:원)', '출생년\n순위', '비고']
for kwd, j in zip(sub, range(1, len(sub)+1)):
    cell = ws.cell(row=4, column=j+1)
    cell.value = kwd
    cell_style(ws, cell)
    cell.fill = PatternFill(fill_type='solid', fgColor=Color('FFC000'))
ws.row_dimensions[4].height = 27.25
# - 자료 정의
employee_code = ["PE-205", "PE-107", "TE-106", "PE-301", "TE-103", "PE-202", "TE-208", "TE-304"]
names = ["김지은", "노승일", "김선정", "배현진", "박성호", "서은하", "장근오", "김재국"]
personnel_transfer = ["재무관리부", "배송부", "배송부", "재무관리부", "배송부", "식료사업부", "식료사업부", "식료사업부"]
division = ["복직", "이동", "채용", "이동", "이동", "이동", "채용", "채용"]
term = [4, 11, 1, 12, 5, 14, 3, 1]
year = [1983, 1979, 1991, 1978, 1980, 1972, 1993, 1985]
pay = [2257000, 4926000, 1886000, 5236000, 2386000, 4436000, 2350000, 1786000]
raw_datas = {
    "사원코드": employee_code,
    "이름": names,
    "발령부서": personnel_transfer,
    "발령구분": division,
    "근속기간": term,
    "출생년": year,
    "급여": pay
}
# - 셀에 기본 데이터 입력
key_list = list(raw_datas.keys())
for row in range(5, 15):
    idx = 0
    for col in range(2, 11):
        try:
            key = key_list[idx]
            value = raw_datas[key][row-5]
        except IndexError:
            value = None
        cell = ws.cell(row=row, column=col)
        if value:
            cell.value = value
        if idx <= 3:
            cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(name='굴림', size=11)
        idx += 1

# 셀 병합
# - 데이터 입력 및 처리
ws["B13"] = "최저 급여(단위:원)"
ws["B14"] = "제무관리부 급여(단위: 원) 평균"
ws["G13"] = "발령구분이 복직인 사원수"
cell_style(ws, ws["G13"])
# - 기타 데이터 입력
g14 = ws["G14"]
i14 = ws["I14"]
g14.value = "사원코드"
i14.value = "근속기간"
cell_style(ws, i14)

# 표 테두리 설정
# - 선 정의
bd_thin = Side(border_style='thin')
bd_thick = Side(border_style='medium')
THIN_BORDER = Border(bd_thin, bd_thin, bd_thin, bd_thin)
THICK_BORDER = Border(
    top=bd_thick, bottom=bd_thick, left=bd_thick, right=bd_thick
)
THICK_Top = Border(
    top=bd_thick, bottom=bd_thin, left=bd_thin, right=bd_thin
)
THICK_Bottom = Border(
    top=bd_thin, bottom=bd_thick, left=bd_thin, right=bd_thin
)
THICK_Left = Border(
    top=bd_thin, bottom=bd_thin, left=bd_thick, right=bd_thin
)
THICK_Right = Border(
    top=bd_thin, bottom=bd_thin, left=bd_thin, right=bd_thick
)
borderTopLeft = Border(
    top=bd_thick, bottom=bd_thin, left=bd_thick, right=bd_thin
)
borderTopRight = Border(
    top=bd_thick, bottom=bd_thin, left=bd_thin, right=bd_thick
)
borderBottomLeft = Border(
    top=bd_thin, bottom=bd_thick, left=bd_thick, right=bd_thin
)
borderBottomRight = Border(
    top=bd_thin, bottom=bd_thick, left=bd_thin, right=bd_thick
)
# - 선 지정 범위 정의
rowTop = 4
rowBot = 14
colLeft = 2
colRight = 10
rows = range(rowTop, rowBot+1)
cols = range(colLeft, colRight+1)
start_cell = chr(64 + rowTop)
end_cell = chr(64 + colRight)
# - 범위 내 셀들을 ws.cell 지정 방식으로 지정해 기본 테두리 설정
for row in rows:
    for col in cols:
        ws.cell(row, col).border = THIN_BORDER
# - 바깥쪽 테두리 굵게 설정
for row in rows:
    for col in cols:
        ws.cell(rowTop, col).border = THICK_Top
        ws.cell(rowBot, col).border = THICK_Bottom
        ws.cell(row, colLeft).border = THICK_Left
        ws.cell(row, colRight).border = THICK_Right
start_cell = "B"
ws[f'{start_cell}{rowTop}'].border = borderTopLeft
ws[f'{start_cell}{rowBot}'].border = borderBottomLeft
ws[f'{end_cell}{rowTop}'].border = borderTopRight
ws[f'{end_cell}{rowBot}'].border = borderBottomRight
# - 특정 행 굵은 바깥 테두리 설정
for rng in ws["B4:J4"]:
    for idx, cell in enumerate(rng):
        if idx == 0:
            cell.border = Border(top=bd_thick, bottom=bd_thick, left=bd_thick, right=bd_thin)
        elif idx == len(rng)-1:
            cell.border = Border(top=bd_thick, bottom=bd_thick, left=bd_thin, right=bd_thick)
        else:
            cell.border = Border(top=bd_thick, bottom=bd_thick, left=bd_thin, right=bd_thin)
# - 서식 복사를 활용해 테두리 설정
for col in cols:
    about_cell = ws.cell(14, col)                # 스타일을 가져올 셀
    copy_cell = ws.cell(12, col)                 # 스타일 복사가 수행될 셀
    copy_cell._style = copy(about_cell._style)   # 스타일 전체 복사
# - 대각선 테두리 그리기
f13 = ws["F13"]
f13.border = Border(top=bd_thick, bottom=bd_thick, left=bd_thin, right=bd_thin, diagonalUp=True, diagonalDown=True, diagonal=Side(border_style="thin"))

# 기타 셀 배경색 설정
# - 셀 병합 수행
ws.merge_cells("B13:D13")
ws.merge_cells("B14:D14")
ws.merge_cells("F13:F14")
ws.merge_cells("G13:I13")
g14.fill = PatternFill(fill_type='solid', fgColor=Color('FFC000'))
cell_style(ws, g14)
i14.fill = PatternFill(fill_type='solid', fgColor=Color('FFC000'))

# 셀 서식 적용
# - 숫자 뒤에 '년' 표시
for rng in ws["F5:F12"]:
    for cell in  rng:
        cell.number_format = '#,##0"년"'
# - 세 자리 단위 콤마
for rng in ws["H5:H12"]:
    for cell in  rng:
        cell.number_format = '0,000'

# 높이/너비 설정
# - 행 높이 설정
for row in range(1, 4):
    ws.row_dimensions[row].height = 22.5
# - 열 너비 설정
width_list = [10.63, 9.63, 13.13, 11.88, 11.88, 11.88, 13, 11, 12]
idx = 0
for col in range(2, 11):
    ws.column_dimensions[get_column_letter(col)].width = width_list[idx]
    idx += 1

# 「H5:H12」영역에 대해 급여로 이름 정의
# Refer.
# -- https://openpyxl.readthedocs.io/en/stable/defined_names.html
# -- https://stackoverflow.com/questions/60047850/python-openpyxl-package-defined-names-does-not-recognize-named-ranges
# -- https://pythoninoffice.com/how-to-work-with-excel-named-range-in-python/
new_range = DefinedName('급여', attr_text='제1작업!$H$5:$H$12')
wb.defined_names.append(new_range)

# 데이터 유효성 검사
# - 유효성 검사를 이용하여 H14셀에 사원코드 영역(B5:B12)
dv = DataValidation(type="list", formula1="=$B$5:$B$12") # allow_blank=False
ws.add_data_validation(dv)
dv.add("H14")
ws["H14"].value = ws["B5"].value
ws["H14"].alignment = Alignment(horizontal='center')

# 함수 문제 풀이
# - (1) 출생년 순위: 출생년 컬럼 기준 오름차순 순위 + '위'
func = "RANK"
for row in range(5, 13):
    formula_range = f"G{row}, $G$5:$G$12, 1"
    cell = ws["I"][row-1]
    cell.value = f'={func}({formula_range})&"위"'
    cell.alignment = Alignment(horizontal='right')
# - (2) 비고: 사원코드 기준 앞 두 글자가 PE면 정규직 그 외에는 계약직
for row in range(5, 13):
    formula_range = f"G{row}, $G$5:$G$12, 1"
    cell = ws["J"][row-1]
    cell.value = f'=IF(LEFT(B{row},2)="PE","정규직", "계약직")'
    cell.alignment = Alignment(horizontal='center')
# - (3) 최저 급여 (단위: 원): 정의된 이름(급여) 이용
cell = ws["E13"]
cell.value = "=MIN(급여)"
cell.alignment = Alignment(horizontal='right')
cell.number_format = '0,000'
# - (4) 재무관리부 급여(단위:원) 평균: 조건은 입력 데이터를 이용, 반올림하여 만 단위까지
cell = ws["E14"]
cell.value = "=ROUND(DAVERAGE(B4:H12,7,D4:D5),-4)"
cell.alignment = Alignment(horizontal='right')
cell.number_format = '0,000'
# - (5) 발령구분이 복직인 사원수: 조건은 입력 데이터 이용
cell = ws["J13"]
cell.value = "=DCOUNTA(B4:H12,4,E4:E5)"
cell.alignment = Alignment(horizontal='right')
# - (6) 근속기간: 「근속기간」셀에서 선택한 사원코드에 대한 근속기간
cell = ws["J14"]
cell.value = "=VLOOKUP(H14,B5:H12,5,0)"
cell.alignment = Alignment(horizontal='right')

# 조건부 서식: 수식을 이용하여 급여 단위(단위:원)가 4,000,000 이상인 행 전체에 '글꼴: 파랑, 굵게'
# - 일반 조건문을 사용한 방법
conditional_font = Font(color='0070C0', bold=True, name='굴림', size=11)
'''
for row in ws["H5:H12"]:
    stat = False
    for cell in row:
        if isinstance(cell.value, int) and cell.value >= 4000000:
            stat = True
        if stat:
            for cell in ws[cell.row]:
                cell.font = conditional_font
'''
# - 조건부 서식에서 새 규칙을 추가해 사용하는 수식을 이용한 방법
# - refer. https://openpyxl.readthedocs.io/en/stable/formatting.html
dxf = DifferentialStyle(font=conditional_font)
rule = Rule(type='expression', formula=['=$H5>=4000000'], dxf=dxf)
ws.conditional_formatting.add('$B$5:$J$12', rule)

# 제목 도형 삽입
# - openpyxl에서 도형 그리기 기능은 제공하지 않는 것으로 확인
# - refer. 도형 미지원 참고 자료
# -- https://openpyxl.readthedocs.io/en/2.4/api/openpyxl.drawing.shapes.html
# "You are not reading the most recent version of this documentation. 2.5.14 is the latest version available."
# -- https://stackoverflow.com/questions/48714562/python-openpyxl-how-to-insert-a-shape-for-text
# -- https://foss.heptapod.net/openpyxl/openpyxl/-/issues/1488

# 이미지 삽입
img = Image('결재.png')
ws.add_image(img, 'H1')

# 완료 데이터 저장
wb.save("result-2201010B-openpyxl_part1.xlsx")
