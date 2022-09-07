"""
220110_B_works1.py
- Post 14. 응용 사례 - ITQ 엑셀 문제 풀이
[22년 01월 10일 기출문제 복원본 풀이]
- 문제 출처 : https://www.comcbt.com/xe/itqe/5563038
"""

# 제1작업
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.styles import PatternFill, Color
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting import Rule
from openpyxl.styles.differential import DifferentialStyle
from copy import copy


""" 제1작업 """
# 새 워크북 생성 및 시트 이름 변경
wb = Workbook()
ws = wb.active
ws.title = "제1작업"

# A열 너비 조정
ws.column_dimensions["A"].width = 1

# 셀 기본 스타일 지정 함수
def cell_style(workseet, cell):
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.font = Font(name='굴림', size=11)
    return workseet

# 데이터 입력
# - 4행 데이터 입력
sub = ['사원코드', '이름', '발령부서', '발령구분', '근속기간', '출생년', '급여\n(단위:원)', '출생년\n순위', '비고']
for kwd, j in zip(sub, range(1, len(sub)+1)):
    cell = ws.cell(row=4, column=j+1)
    cell.value = kwd
    cell_style(ws, cell)
    cell.fill = PatternFill(fill_type='solid', fgColor=Color('FFC000'))
ws.row_dimensions[4].height = 27.25
# - 자료 정의
employee_code = ["PE-205", "PE-107", "TE-106", "PE-301", "TE-103", "PE-202", "TE-208", "TE-304"]
names = ["김지은", "노승일", "김선정", "배현진", "박성호", "서은하", "장근오", "김재국"]
personnel_transfer = ["재무관리부", "배송부", "배송부", "재무관리부", "배송부", "식료사업부", "식료사업부", "식료사업부"]
division = ["복직", "이동", "채용", "이동", "이동", "이동", "채용", "채용"]
term = [4, 11, 1, 12, 5, 14, 3, 1]
year = [1983, 1979, 1991, 1978, 1980, 1972, 1993, 1985]
pay = [2257000, 4926000, 1886000, 5236000, 2386000, 4436000, 2350000, 1786000]
raw_datas = {
    "사원코드": employee_code,
    "이름": names,
    "발령부서": personnel_transfer,
    "발령구분": division,
    "근속기간": term,
    "출생년": year,
    "급여": pay
}
# - 셀에 기본 데이터 입력
key_list = list(raw_datas.keys())
for row in range(5, 15):
    idx = 0
    for col in range(2, 11):
        try:
            key = key_list[idx]
            value = raw_datas[key][row-5]
        except IndexError:
            value = None
        cell = ws.cell(row=row, column=col)
        if value:
            cell.value = value
        if idx <= 3:
            cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(name='굴림', size=11)
        idx += 1

# 셀 병합
# - 데이터 입력 및 처리
ws["B13"] = "최저 급여(단위:원)"
ws["B14"] = "제무관리부 급여(단위: 원) 평균"
ws["G13"] = "발령구분이 복직인 사원수"
cell_style(ws, ws["G13"])
# - 기타 데이터 입력
g14 = ws["G14"]
i14 = ws["I14"]
g14.value = "사원코드"
i14.value = "근속기간"
cell_style(ws, i14)

# 표 테두리 설정
# - 선 정의
bd_thin = Side(border_style='thin')
bd_thick = Side(border_style='medium')
THIN_BORDER = Border(bd_thin, bd_thin, bd_thin, bd_thin)
THICK_BORDER = Border(
    top=bd_thick, bottom=bd_thick, left=bd_thick, right=bd_thick
)
THICK_Top = Border(
    top=bd_thick, bottom=bd_thin, left=bd_thin, right=bd_thin
)
THICK_Bottom = Border(
    top=bd_thin, bottom=bd_thick, left=bd_thin, right=bd_thin
)
THICK_Left = Border(
    top=bd_thin, bottom=bd_thin, left=bd_thick, right=bd_thin
)
THICK_Right = Border(
    top=bd_thin, bottom=bd_thin, left=bd_thin, right=bd_thick
)
borderTopLeft = Border(
    top=bd_thick, bottom=bd_thin, left=bd_thick, right=bd_thin
)
borderTopRight = Border(
    top=bd_thick, bottom=bd_thin, left=bd_thin, right=bd_thick
)
borderBottomLeft = Border(
    top=bd_thin, bottom=bd_thick, left=bd_thick, right=bd_thin
)
borderBottomRight = Border(
    top=bd_thin, bottom=bd_thick, left=bd_thin, right=bd_thick
)
# - 선 지정 범위 정의
rowTop = 4
rowBot = 14
colLeft = 2
colRight = 10
rows = range(rowTop, rowBot+1)
cols = range(colLeft, colRight+1)
start_cell = chr(64 + rowTop)
end_cell = chr(64 + colRight)
# - 범위 내 셀들을 ws.cell 지정 방식으로 지정해 기본 테두리 설정
for row in rows:
    for col in cols:
        ws.cell(row, col).border = THIN_BORDER
# - 바깥쪽 테두리 굵게 설정
for row in rows:
    for col in cols:
        ws.cell(rowTop, col).border = THICK_Top
        ws.cell(rowBot, col).border = THICK_Bottom
        ws.cell(row, colLeft).border = THICK_Left
        ws.cell(row, colRight).border = THICK_Right
start_cell = "B"
ws[f'{start_cell}{rowTop}'].border = borderTopLeft
ws[f'{start_cell}{rowBot}'].border = borderBottomLeft
ws[f'{end_cell}{rowTop}'].border = borderTopRight
ws[f'{end_cell}{rowBot}'].border = borderBottomRight
# - 특정 행 굵은 바깥 테두리 설정
for rng in ws["B4:J4"]:
    for idx, cell in enumerate(rng):
        if idx == 0:
            cell.border = Border(top=bd_thick, bottom=bd_thick, left=bd_thick, right=bd_thin)
        elif idx == len(rng)-1:
            cell.border = Border(top=bd_thick, bottom=bd_thick, left=bd_thin, right=bd_thick)
        else:
            cell.border = Border(top=bd_thick, bottom=bd_thick, left=bd_thin, right=bd_thin)
# - 서식 복사를 활용해 테두리 설정
for col in cols:
    about_cell = ws.cell(14, col)                # 스타일을 가져올 셀
    copy_cell = ws.cell(12, col)                 # 스타일 복사가 수행될 셀
    copy_cell._style = copy(about_cell._style)   # 스타일 전체 복사
# - 대각선 테두리 그리기
f13 = ws["F13"]
f13.border = Border(top=bd_thick, bottom=bd_thick, left=bd_thin, right=bd_thin, diagonalUp=True, diagonalDown=True, diagonal=Side(border_style="thin"))

# 기타 셀 배경색 설정
# - 셀 병합 수행
ws.merge_cells("B13:D13")
ws.merge_cells("B14:D14")
ws.merge_cells("F13:F14")
ws.merge_cells("G13:I13")
g14.fill = PatternFill(fill_type='solid', fgColor=Color('FFC000'))
cell_style(ws, g14)
i14.fill = PatternFill(fill_type='solid', fgColor=Color('FFC000'))

# 셀 서식 적용
# - 숫자 뒤에 '년' 표시
for rng in ws["F5:F12"]:
    for cell in  rng:
        cell.number_format = '#,##0"년"'
# - 세 자리 단위 콤마
for rng in ws["H5:H12"]:
    for cell in  rng:
        cell.number_format = '0,000'

# 높이/너비 설정
# - 행 높이 설정
for row in range(1, 4):
    ws.row_dimensions[row].height = 22.5
# - 열 너비 설정
width_list = [10.63, 9.63, 13.13, 11.88, 11.88, 11.88, 13, 11, 12]
idx = 0
for col in range(2, 11):
    ws.column_dimensions[get_column_letter(col)].width = width_list[idx]
    idx += 1

# 「H5:H12」영역에 대해 급여로 이름 정의
# Refer.
# -- https://openpyxl.readthedocs.io/en/stable/defined_names.html
# -- https://stackoverflow.com/questions/60047850/python-openpyxl-package-defined-names-does-not-recognize-named-ranges
# -- https://pythoninoffice.com/how-to-work-with-excel-named-range-in-python/
new_range = DefinedName('급여', attr_text='제1작업!$H$5:$H$12')
wb.defined_names.append(new_range)

# 데이터 유효성 검사
# - 유효성 검사를 이용하여 H14셀에 사원코드 영역(B5:B12)
dv = DataValidation(type="list", formula1="=$B$5:$B$12") # allow_blank=False
ws.add_data_validation(dv)
dv.add("H14")
ws["H14"].value = ws["B5"].value
ws["H14"].alignment = Alignment(horizontal='center')

# 함수 문제 풀이
# - (1) 출생년 순위: 출생년 컬럼 기준 오름차순 순위 + '위'
func = "RANK"
for row in range(5, 13):
    formula_range = f"G{row}, $G$5:$G$12, 1"
    cell = ws["I"][row-1]
    cell.value = f'={func}({formula_range})&"위"'
    cell.alignment = Alignment(horizontal='right')
# - (2) 비고: 사원코드 기준 앞 두 글자가 PE면 정규직 그 외에는 계약직
for row in range(5, 13):
    formula_range = f"G{row}, $G$5:$G$12, 1"
    cell = ws["J"][row-1]
    cell.value = f'=IF(LEFT(B{row},2)="PE","정규직", "계약직")'
    cell.alignment = Alignment(horizontal='center')
# - (3) 최저 급여 (단위: 원): 정의된 이름(급여) 이용
cell = ws["E13"]
cell.value = "=MIN(급여)"
cell.alignment = Alignment(horizontal='right')
cell.number_format = '0,000'
# - (4) 재무관리부 급여(단위:원) 평균: 조건은 입력 데이터를 이용, 반올림하여 만 단위까지
cell = ws["E14"]
cell.value = "=ROUND(DAVERAGE(B4:H12,7,D4:D5),-4)"
cell.alignment = Alignment(horizontal='right')
cell.number_format = '0,000'
# - (5) 발령구분이 복직인 사원수: 조건은 입력 데이터 이용
cell = ws["J13"]
cell.value = "=DCOUNTA(B4:H12,4,E4:E5)"
cell.alignment = Alignment(horizontal='right')
# - (6) 근속기간: 「근속기간」셀에서 선택한 사원코드에 대한 근속기간
cell = ws["J14"]
cell.value = "=VLOOKUP(H14,B5:H12,5,0)"
cell.alignment = Alignment(horizontal='right')

# 조건부 서식: 수식을 이용하여 급여 단위(단위:원)가 4,000,000 이상인 행 전체에 '글꼴: 파랑, 굵게'
# - 일반 조건문을 사용한 방법
conditional_font = Font(color='0070C0', bold=True, name='굴림', size=11)
# - 조건부 서식에서 새 규칙을 추가해 사용하는 수식을 이용한 방법
# - refer. https://openpyxl.readthedocs.io/en/stable/formatting.html
dxf = DifferentialStyle(font=conditional_font)
rule = Rule(type='expression', formula=['=$H5>=4000000'], dxf=dxf)
ws.conditional_formatting.add('$B$5:$J$12', rule)

# 제목 도형 삽입
# - openpyxl에서 도형 그리기 기능은 제공하지 않는 것으로 확인
# - refer. 도형 미지원 참고 자료
# -- https://openpyxl.readthedocs.io/en/2.4/api/openpyxl.drawing.shapes.html
# "You are not reading the most recent version of this documentation. 2.5.14 is the latest version available."
# -- https://stackoverflow.com/questions/48714562/python-openpyxl-how-to-insert-a-shape-for-text
# -- https://foss.heptapod.net/openpyxl/openpyxl/-/issues/1488

# 이미지 삽입
img = Image('결재.png')
ws.add_image(img, 'H1')


""" 제2작업 """
# 제2작업
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd

ws2 = wb.create_sheet("제2작업")

# A열 너비 조정
ws2.column_dimensions["A"].width = 1

# 셀 기본 스타일 지정 함수
def cell_style(workseet, cell, align=True, thin=True):
    cell.font = Font(name='굴림', size=11)
    if align:
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    if thin:
        cell.border = THIN_BORDER
    return workseet

# 기존 데이터를 다른 시트로 복사
min_row, max_row = 4, 12
min_col, max_col = 2, 8
# copying the cell values from source
# excel file to destination excel file
for i in range(min_row, max_row+1):
    for j in range (min_col, max_col+1):
        # reading cell value from source excel file
        origin_cell = ws.cell(row=i, column=j)
        # writing the read value to destination excel file
        cell = ws2.cell(row=i-2, column=j)
        cell.value = origin_cell.value
        cell._style = copy(origin_cell._style)

# 조건부 서식: 수식을 이용하여 급여 단위(단위:원)가 4,000,000 이상인 행 전체에 '글꼴: 파랑, 굵게'
# - refer. https://openpyxl.readthedocs.io/en/stable/formatting.html
dxf = DifferentialStyle(font=conditional_font)
rule = Rule(type='expression', formula=['=$H3>=4000000'], dxf=dxf)
ws2.conditional_formatting.add('$B$3:$H$10', rule)

# 테두리 정의
bd_thin = Side(border_style='thin')
THIN_BORDER = Border(bd_thin, bd_thin, bd_thin, bd_thin)

# B11, H11 - 급여 전체 평균
b11 = ws2["B11"]
b11.value = "급여(단위:원) 전체 평균"
cell_style(ws2, b11)
ws2.merge_cells("B11:G11")
h11 = ws2["H11"]
h11.value = "=AVERAGE(H3:H10)"
h11.number_format = '0,000'
cell_style(ws2, h11)

# 목표값 찾기
# - 수식 셀: H11
# - 찾는 값: 3200000
# - 값을 바꿀 셀: $H$3 (원본 값 변경)
# *** 목표값 칮기 기능 자체는 openpyxl의 미지원 기능으로 판단 ***
# 대안: 다음으로 해당 기능 수행 결과 대체 가능
# 변경대상값 = 기존 평균값에서 변경할 평균 값 * 총 자료수 - 대상값을 제외한 합
# 23006000 + x / 8 = 3200000 => x = 3200000 * 8 - 23006000
nums = 0
for row in ws2["H4:H10"]:
    for cell in row:
        nums += cell.value
ws2["H11"].value = 3200000
ws2["H3"].value = 3200000 * 8 - nums

# 고급필터
# 발령부서가 '배송부'이거나, 근속기간이 '2'이하인 자료 이름, 발령구분, 근속기간, 급여 데이터만 추출
# - 조건 범위: B14 셀부터 입력
# - 복사 위치: B18 셀부터 나타나도록
# *** 고급필터 기능 자체는 openpyxl의 미지원 기능으로 판단 ***

# 대안: 파이썬 판다스 필터 사용
# - filter 조건
# -> ws2[B14], ws2[B15]
cell = ws2["B14"]
cell.value = "발령부서"
cell._style = copy(ws2["D2"]._style)
cell = ws2["B15"]
cell.value = "배송부"
cell.alignment = Alignment(horizontal='left')
cell_style(ws2, cell, False, False)
# -> ws2[C14], ws2[C15]
cell = ws2["C14"]
cell.value = "근속기간"
cell._style = copy(ws2["F2"]._style)
cell = ws2["C16"]
cell.value = "<=2"
cell.alignment = Alignment(horizontal='left')
cell_style(ws2, cell, False, False)
# - 필터 데이터 적용
df = pd.DataFrame(ws2.values)
df = df.drop(0, axis=1)
df.columns = df.iloc[1, :]
df = df[2:-1]
mask1 = (df.발령부서 == "배송부") | (df.근속기간 <= 2)
df_filter = df.loc[mask1,:]
df_filter = df_filter[["이름", "발령구분", "근속기간", "급여\n(단위:원)"]]
for idx, r in enumerate(dataframe_to_rows(df_filter, index=True, header=True)):
    # 빈 행 추가
    if idx == 0:
        ws2.append([None])
        ws2.append(r)
    elif r == [None]:
        continue
    else:
        ws2.append(r)


""" 제4작업 """
# 제4작업
# - chart generate
from openpyxl.chart import LineChart, BarChart, Reference, Series
# - chart font
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties, Font
# - chart color set
from openpyxl.drawing.line import LineProperties
from openpyxl.drawing.fill import PatternFillProperties, ColorChoice
# - chart labels
from openpyxl.chart.label import DataLabelList
# - chart background color setting
from openpyxl.chart.shapes import GraphicalProperties

ws4 = wb.create_chartsheet("제4작업")

# refer
refer_names = "제1작업!$C$5:$C$12"
refer_money = "제1작업!$H$5:$H$12"
refer_time =  "제1작업!$F$5:$F$12"

# bar chart
chart_bar = BarChart()
refer_name_string = Reference(ws4, range_string=refer_names)
refer_name_position = Reference(ws4, min_col=3, min_row=5, max_row=12)
refer_money_string = Reference(ws4, range_string=refer_money)
refer_money_position = Reference(ws4,
    min_col=6,
    min_row=5,
    max_row=12
)

v1 = Series(refer_name_string)
data_money_string = (Series(refer_money_string, title="급여(단위: 원)"),)
data_money_position = (Series(refer_money_position),)

chart_bar.series = data_money_string
chart_bar.set_categories(refer_name_string)
chart_bar.title = '배송부 및 식료사업부 급여 현황'
chart_bar.y_axis.number_format = '#,##'

# data labels
chart_bar.dataLabels = DataLabelList()
chart_bar.dataLabels.showVal = True

# line chart
chart_line = LineChart()
refer3 = Reference(ws4, range_string=refer_time)

data2 = Series(refer3, title="근속기간")
chart_line.series = (data2,)
chart_line.set_categories(refer_name_string)
chart_line.y_axis.axId = 200
chart_line.y_axis.number_format = '#,##0"년"'
chart_line.y_axis.scaling.min = 0
chart_line.y_axis.scaling.max = 15
chart_line.y_axis.majorUnit = 3

# Marker line chart
s1 = chart_line.series[0]
s1.marker.symbol = "diamond"
s1.marker.size = 10

# chart style addtional
# Style chart: X and Y axes numbers
def set_chart_title_size(chart, size=1100):
    cp = CharacterProperties(sz=size)
    paraprops = ParagraphProperties(defRPr=cp)
    # paraprops.defRPr = CharacterProperties(latin=font, sz=size)
    for para in chart.title.tx.rich.paragraphs:
        para.pPr=paraprops


# Font
# refer. https://openpyxl.readthedocs.io/en/stable/styles.html
# refer. https://stackoverflow.com/questions/39054631/openpyxl-chage-font-size-of-title-y-axis-title
font = Font(typeface='굴림')
# - 11 point size
size = 1100
# - Not bold
cp = CharacterProperties(latin=font, sz=size, b=False)
pp = ParagraphProperties(defRPr=cp)
rtp = RichText(p=[Paragraph(pPr=pp, endParaRPr=cp)])
# - y축, txPr=textProperties
chart_bar.y_axis.txPr = rtp
chart_line.y_axis.txPr = rtp
# - x 축
chart_line.x_axis.txPr = rtp
# - 범례
chart_line.legend.txPr = rtp
# - 제목
set_chart_title_size(chart_bar, size=2000)
# - 데이터 라벨
chart_bar.dataLabels.txPr = rtp

# 차트 배경색
props = GraphicalProperties(solidFill="f2f2f2")
# print(help(GraphicalProperties()))
chart_bar.plot_area.graphicalProperties = props

# 기본 스타일 시트를 이용한 색상 설정
# chart_line.style = 10

# 색상 설정
# refer. https://openpyxl.readthedocs.io/en/stable/charts/pattern.html
# refer. https://openpyxl.readthedocs.io/en/stable/api/openpyxl.drawing.colors.html
# - 막대 그래프 데이터 계열
series = chart_bar.series[0]
fill =  PatternFillProperties(prst="pct5")
# fill.foreground = ColorChoice(prstClr="red")
fill.foreground = ColorChoice(srgbClr="ED7D31")
fill.background = ColorChoice(srgbClr="ED7D31")
series.graphicalProperties.pattFill = fill
# - 꺾은선 그래프 선
# refer. https://stackoverflow.com/questions/34500606/how-to-set-line-color-of-openpyxl-scatterchart
series = chart_line.series[0]
lineProp = LineProperties(solidFill=ColorChoice(srgbClr='4472C4'))
series.graphicalProperties.line = lineProp
# - 마커 색상 설정
# refer. https://www.shibutan-bloomers.com/python_libraly_openpyxl-9_en/5629/
series = chart_line.series[0]
series.marker.graphicalProperties.solidFill = '4472C4'       # Set marker fill color
series.marker.graphicalProperties.line.solidFill = '4472C4'  # Set marker border color

# 꺾은선 그래프의 눈금선 제거 설정
chart_line.y_axis.majorGridlines = None


# second axis
# refer. https://openpyxl.readthedocs.io/en/stable/charts/secondary.html
chart_line.y_axis.crosses = "max"
chart_bar += chart_line

# 범례 위치 설정
# refer. https://openpyxl.readthedocs.io/en/latest/charts/chart_layout.html
chart_bar.legend.position = 'b'

# Add chart to sheet
ws4.add_chart(chart_bar)

# 완료 데이터 저장
wb.save("result-2201010B-openpyxl.xlsx")
