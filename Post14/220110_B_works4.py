# 필요 라이브러리 호출
# - load workbook
from openpyxl import load_workbook
# - chart generate
from openpyxl.chart import LineChart, BarChart, Reference, Series
# - chart font
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties, Font
# - chart color set
from openpyxl.drawing.line import LineProperties
from openpyxl.drawing.fill import PatternFillProperties, ColorChoice
# - chart labels
from openpyxl.chart.label import DataLabelList
# - chart background color setting
from openpyxl.chart.shapes import GraphicalProperties
# - legend layout setting
from openpyxl.chart.layout import Layout, ManualLayout

# 제1작업 시트 불러오기
wb = load_workbook('result-2201010B-openpyxl_part1.xlsx')
ws = wb["제1작업"]

# 제4작업 차트 시트 생성
ws4 = wb.create_chartsheet("제4작업")

# refer
refer_names = "제1작업!$C$5:$C$12"
refer_money = "제1작업!$H$5:$H$12"
refer_time =  "제1작업!$F$5:$F$12"


""" bar chart """
chart_bar = BarChart()
refer_name_string = Reference(ws4, range_string=refer_names)
refer_name_position = Reference(ws4, min_col=3, min_row=5, max_row=12)
refer_money_string = Reference(ws4, range_string=refer_money)
refer_money_position = Reference(ws4,
    min_col=6,
    min_row=5,
    max_row=12
)

v1 = Series(refer_name_string)
data_money_string = (Series(refer_money_string, title="급여(단위: 원)"),)
data_money_position = (Series(refer_money_position),)

chart_bar.series = data_money_string
chart_bar.set_categories(refer_name_string)
chart_bar.title = '배송부 및 식료사업부 급여 현황'
chart_bar.y_axis.number_format = '#,##'

# data labels
chart_bar.dataLabels = DataLabelList()
chart_bar.dataLabels.showVal = True


""" line chart """
chart_line = LineChart()
refer3 = Reference(ws4, range_string=refer_time)

data2 = Series(refer3, title="근속기간")
chart_line.series = (data2,)
chart_line.set_categories(refer_name_string)
chart_line.y_axis.axId = 200
chart_line.y_axis.number_format = '#,##0"년"'
chart_line.y_axis.scaling.min = 0
chart_line.y_axis.scaling.max = 15
chart_line.y_axis.majorUnit = 3

# Marker line chart
s1 = chart_line.series[0]
s1.marker.symbol = "diamond"
s1.marker.size = 10

""" chart style addtional """
# Style chart: X and Y axes numbers
def set_chart_title_size(chart, size=1100):
    cp = CharacterProperties(sz=size)
    paraprops = ParagraphProperties(defRPr=cp)
    # paraprops.defRPr = CharacterProperties(latin=font, sz=size)
    for para in chart.title.tx.rich.paragraphs:
        para.pPr=paraprops


# Font
# refer. https://openpyxl.readthedocs.io/en/stable/styles.html
# refer. https://stackoverflow.com/questions/39054631/openpyxl-chage-font-size-of-title-y-axis-title
font = Font(typeface='굴림')
# - 11 point size
size = 1100
# - Not bold
cp = CharacterProperties(latin=font, sz=size, b=False)
pp = ParagraphProperties(defRPr=cp)
rtp = RichText(p=[Paragraph(pPr=pp, endParaRPr=cp)])
# - y축, txPr=textProperties
chart_bar.y_axis.txPr = rtp
chart_line.y_axis.txPr = rtp
# - x 축
chart_line.x_axis.txPr = rtp
# - 범례
chart_line.legend.txPr = rtp
# - 제목
set_chart_title_size(chart_bar, size=2000)
# - 데이터 라벨
chart_bar.dataLabels.txPr = rtp

# 차트 배경색
props = GraphicalProperties(solidFill="f2f2f2")
# print(help(GraphicalProperties()))
chart_bar.plot_area.graphicalProperties = props

# 기본 스타일 시트를 이용한 색상 설정
# chart_line.style = 10

# 색상 설정
# refer. https://openpyxl.readthedocs.io/en/stable/charts/pattern.html
# refer. https://openpyxl.readthedocs.io/en/stable/api/openpyxl.drawing.colors.html
# - 막대 그래프 데이터 계열
series = chart_bar.series[0]
fill =  PatternFillProperties(prst="pct5")
# fill.foreground = ColorChoice(prstClr="red")
fill.foreground = ColorChoice(srgbClr="ED7D31")
fill.background = ColorChoice(srgbClr="ED7D31")
series.graphicalProperties.pattFill = fill
# - 꺾은선 그래프 선
# refer. https://stackoverflow.com/questions/34500606/how-to-set-line-color-of-openpyxl-scatterchart
series = chart_line.series[0]
lineProp = LineProperties(solidFill=ColorChoice(srgbClr='4472C4'))
series.graphicalProperties.line = lineProp
# - 마커 색상 설정
# refer. https://www.shibutan-bloomers.com/python_libraly_openpyxl-9_en/5629/
series = chart_line.series[0]
series.marker.graphicalProperties.solidFill = '4472C4'       # Set marker fill color
series.marker.graphicalProperties.line.solidFill = '4472C4'  # Set marker border color

# 꺾은선 그래프의 눈금선 제거 설정
chart_line.y_axis.majorGridlines = None


""" Adding a second axis """
# refer. https://openpyxl.readthedocs.io/en/stable/charts/secondary.html
chart_line.y_axis.crosses = "max"
chart_bar += chart_line

# 범례 위치 설정
# refer. https://openpyxl.readthedocs.io/en/latest/charts/chart_layout.html
chart_bar.legend.position = 'b'

'''
chart_line.legend.position = 'tr'
chart_bar.legend.position = "tr"
chart_bar.legend.layout = Layout(
    manualLayout=ManualLayout(
        yMode='edge',
        xMode='edge',
        x=0, y=0.9,
        h=0.1, w=0.5
    )
)
'''

# Add chart to sheet
ws4.add_chart(chart_bar)

# 완료 데이터 저장
wb.save("result-2201010B-openpyxl_part4.xlsx")
